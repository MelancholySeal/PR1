import openpyxl

workbook = openpyxl.load_workbook('18.xlsx')

sheet = workbook.active

N = sheet.max_row

coins = [[0] * N for _ in range(N)]

for i in range(N):
    for j in range(N):
        coins[i][j] = sheet.cell(row=i + 1, column=j + 1).value

max_coins = [[0] * N for _ in range(N)]
min_coins = [[0] * N for _ in range(N)]

max_coins[0][0] = min_coins[0][0] = coins[0][0]

for i in range(1, N):
    max_coins[i][0] = min_coins[i][0] = max_coins[0][i] = min_coins[0][i] = max_coins[0][i - 1] + coins[0][i]

for i in range(1, N):
    for j in range(1, N):
        max_coins[i][j] = max(max_coins[i - 1][j], max_coins[i][j - 1]) + coins[i][j]
        min_coins[i][j] = min(min_coins[i - 1][j], min_coins[i][j - 1]) + coins[i][j]

max_sum = max_coins[N - 1][N - 1]

min_sum = min_coins[N - 1][N - 1]

workbook.close()

print(max_sum, min_sum)
